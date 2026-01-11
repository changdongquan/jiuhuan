#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析Excel文件中R列图片是否在单元格内
"""
import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, OneCellAnchor
import xlrd
from openpyxl import Workbook
from copy import copy
import xml.etree.ElementTree as ET
import zipfile
import os

def convert_xls_to_xlsx_with_libreoffice(xls_path, xlsx_path):
    """使用LibreOffice将.xls文件转换为.xlsx格式（保留图片）"""
    import subprocess
    import os
    
    try:
        libreoffice_paths = [
            'libreoffice',
            '/Applications/LibreOffice.app/Contents/MacOS/soffice'
        ]
        
        soffice = None
        for path in libreoffice_paths:
            if os.path.exists(path) or subprocess.run(['which', path.split('/')[-1]], 
                                                      capture_output=True).returncode == 0:
                soffice = path
                break
        
        if not soffice:
            print("未找到LibreOffice")
            return False
        
        # 获取输出目录
        output_dir = os.path.dirname(xlsx_path)
        input_file = os.path.abspath(xls_path)
        
        print(f"使用LibreOffice转换文件...")
        print(f"输入: {input_file}")
        print(f"输出目录: {output_dir}")
        
        # LibreOffice转换命令
        cmd = [
            soffice,
            '--headless',
            '--convert-to', 'xlsx',
            '--outdir', output_dir,
            input_file
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode == 0:
            # LibreOffice会生成与输入文件同名的.xlsx文件
            expected_output = os.path.join(output_dir, 
                                          os.path.splitext(os.path.basename(xls_path))[0] + '.xlsx')
            if os.path.exists(expected_output):
                if expected_output != xlsx_path:
                    # 重命名到目标路径
                    os.rename(expected_output, xlsx_path)
                print(f"✓ LibreOffice转换成功: {xlsx_path}")
                return True
            else:
                print(f"转换完成，但未找到输出文件: {expected_output}")
                return False
        else:
            print(f"LibreOffice转换失败:")
            print(result.stderr)
            return False
            
    except subprocess.TimeoutExpired:
        print("转换超时")
        return False
    except Exception as e:
        print(f"转换错误: {e}")
        import traceback
        traceback.print_exc()
        return False

def convert_xls_to_xlsx(xls_path, xlsx_path):
    """将.xls文件转换为.xlsx格式"""
    try:
        print(f"正在转换 {xls_path} 到 {xlsx_path}...")
        
        # 读取xls文件
        xls_workbook = xlrd.open_workbook(xls_path)
        
        # 创建xlsx工作簿
        xlsx_workbook = Workbook()
        xlsx_workbook.remove(xlsx_workbook.active)  # 删除默认工作表
        
        # 复制每个工作表
        for sheet_idx in range(xls_workbook.nsheets):
            xls_sheet = xls_workbook.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_workbook.create_sheet(title=xls_sheet.name)
            
            # 复制单元格数据
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    xlsx_cell = xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1)
                    
                    # 处理不同类型的值
                    cell_type = xls_sheet.cell_type(row_idx, col_idx)
                    if cell_type == xlrd.XL_CELL_DATE:
                        # 处理日期
                        try:
                            date_tuple = xlrd.xldate_as_tuple(cell_value, xls_workbook.datemode)
                            xlsx_cell.value = date_tuple
                        except:
                            xlsx_cell.value = cell_value
                    else:
                        xlsx_cell.value = cell_value
                    
                    # 复制格式（尽可能）
                    xls_cell = xls_sheet.cell(row_idx, col_idx)
                    # 注意：格式复制有限，主要保留数据
        
        xlsx_workbook.save(xlsx_path)
        print(f"转换完成: {xlsx_path}")
        return True
        
    except Exception as e:
        print(f"转换失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def parse_drawing_xml(file_path, sheet_name, column_num):
    """解析drawing XML文件来查找图片位置"""
    try:
        # xlsx文件实际上是一个ZIP文件
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            # 查找drawing文件
            drawing_path = f'xl/drawings/drawing1.xml'
            
            if drawing_path not in zip_ref.namelist():
                # 可能有多个drawing文件，尝试查找
                drawing_files = [f for f in zip_ref.namelist() if 'xl/drawings/drawing' in f and f.endswith('.xml')]
                if not drawing_files:
                    return []
                drawing_path = drawing_files[0]
            
            # 读取drawing XML
            with zip_ref.open(drawing_path) as f:
                tree = ET.parse(f)
                root = tree.getroot()
            
            # 解析命名空间
            namespaces = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            
            images_info = []
            # 查找所有图片锚点
            for two_cell_anchor in root.findall('.//xdr:twoCellAnchor', namespaces):
                from_col = None
                from_row = None
                to_col = None
                to_row = None
                
                # 获取起始位置
                from_elem = two_cell_anchor.find('.//xdr:from', namespaces)
                if from_elem is not None:
                    col_elem = from_elem.find('xdr:col', namespaces)
                    row_elem = from_elem.find('xdr:row', namespaces)
                    if col_elem is not None:
                        from_col = int(col_elem.text) + 1  # Excel从1开始
                    if row_elem is not None:
                        from_row = int(row_elem.text) + 1
                
                # 获取结束位置
                to_elem = two_cell_anchor.find('.//xdr:to', namespaces)
                if to_elem is not None:
                    col_elem = to_elem.find('xdr:col', namespaces)
                    row_elem = to_elem.find('xdr:row', namespaces)
                    if col_elem is not None:
                        to_col = int(col_elem.text) + 1
                    if row_elem is not None:
                        to_row = int(row_elem.text) + 1
                
                if from_col is not None:
                    images_info.append({
                        'from_col': from_col,
                        'from_row': from_row,
                        'to_col': to_col,
                        'to_row': to_row,
                        'in_target_column': (from_col <= column_num <= (to_col or from_col))
                    })
            
            # 也检查oneCellAnchor（单单元格锚点）
            for one_cell_anchor in root.findall('.//xdr:oneCellAnchor', namespaces):
                from_col = None
                from_row = None
                
                from_elem = one_cell_anchor.find('.//xdr:from', namespaces)
                if from_elem is not None:
                    col_elem = from_elem.find('xdr:col', namespaces)
                    row_elem = from_elem.find('xdr:row', namespaces)
                    if col_elem is not None:
                        from_col = int(col_elem.text) + 1
                    if row_elem is not None:
                        from_row = int(row_elem.text) + 1
                
                if from_col is not None:
                    images_info.append({
                        'from_col': from_col,
                        'from_row': from_row,
                        'to_col': None,
                        'to_row': None,
                        'in_target_column': (from_col == column_num)
                    })
            
            return images_info
            
    except Exception as e:
        print(f"  解析drawing XML时出错: {e}")
        return []

def analyze_images_in_xlsx(file_path, column_letter='R'):
    """分析.xlsx文件中指定列的图片位置"""
    try:
        wb = load_workbook(file_path, data_only=False, keep_links=False)
        column_num = column_index_from_string(column_letter)
        
        print(f"\n{'='*70}")
        print(f"分析文件: {file_path}")
        print(f"目标列: {column_letter} (列号: {column_num})")
        print(f"{'='*70}\n")
        
        # 分析每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n工作表: {sheet_name}")
            print(f"最大行数: {ws.max_row}, 最大列数: {ws.max_column}")
            print("-" * 70)
            
            # 检查图片
            images_found = False
            if hasattr(ws, '_images') and ws._images:
                images_found = True
                print(f"\n找到 {len(ws._images)} 个图片对象")
                
                for idx, img in enumerate(ws._images, 1):
                    print(f"\n【图片 {idx}】")
                    print(f"  类型: {type(img).__name__}")
                    
                    # 分析锚点位置
                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        anchor_type = type(anchor).__name__
                        print(f"  锚点类型: {anchor_type}")
                        
                        # TwoCellAnchor: 跨越多个单元格
                        # OneCellAnchor: 单个单元格锚点
                        if isinstance(anchor, TwoCellAnchor):
                            from_col = anchor._from.col + 1  # openpyxl从0开始，Excel从1开始
                            from_row = anchor._from.row + 1
                            to_col = anchor._to.col + 1
                            to_row = anchor._to.row + 1
                            
                            print(f"  起始位置: 列 {get_column_letter(from_col)} (列号: {from_col}), 行 {from_row}")
                            print(f"  结束位置: 列 {get_column_letter(to_col)} (列号: {to_col}), 行 {to_row}")
                            
                            # 检查是否跨越目标列R
                            if from_col <= column_num <= to_col:
                                print(f"  ✓ 图片跨越目标列 {column_letter}")
                                if from_col == column_num and to_col == column_num:
                                    print(f"  ✓ 图片完全在列 {column_letter} 内")
                                else:
                                    print(f"  ⚠ 图片跨越多个列（包括列 {column_letter}）")
                            else:
                                print(f"  ✗ 图片不在列 {column_letter} 范围内")
                        
                        elif isinstance(anchor, OneCellAnchor):
                            col = anchor._from.col + 1
                            row = anchor._from.row + 1
                            
                            print(f"  位置: 列 {get_column_letter(col)} (列号: {col}), 行 {row}")
                            
                            if col == column_num:
                                print(f"  ✓ 图片在列 {column_letter} 内")
                            else:
                                print(f"  ✗ 图片不在列 {column_letter}（在列 {get_column_letter(col)}）")
                        
                        # 显示偏移量信息（如果图片不完全对齐单元格）
                        if hasattr(anchor._from, 'rowOff') and hasattr(anchor._from, 'colOff'):
                            row_off = anchor._from.rowOff if hasattr(anchor._from, 'rowOff') else 0
                            col_off = anchor._from.colOff if hasattr(anchor._from, 'colOff') else 0
                            if row_off != 0 or col_off != 0:
                                print(f"  ⚠ 图片有偏移: 行偏移 {row_off}, 列偏移 {col_off}")
                                print(f"  ⚠ 图片可能不完全对齐单元格")
            
            if not images_found:
                print("\n未找到标准图片对象，尝试解析drawing XML...")
            
            # 检查drawing关系并解析
            if hasattr(ws, '_rels') and ws._rels:
                try:
                    # _rels是RelationshipList，需要遍历
                    drawing_rels = []
                    for rel in ws._rels:
                        target = str(getattr(rel, 'target', ''))
                        if 'drawing' in target.lower():
                            drawing_rels.append(rel)
                    
                    if drawing_rels:
                        print(f"\n找到 {len(drawing_rels)} 个绘图关系，正在解析...")
                        
                        # 解析drawing XML
                        images_info = parse_drawing_xml(file_path, sheet_name, column_num)
                        
                        if images_info:
                            print(f"\n找到 {len(images_info)} 个图片/形状对象：")
                            for idx, img_info in enumerate(images_info, 1):
                                print(f"\n【图片/形状 {idx}】")
                                print(f"  起始位置: 列 {get_column_letter(img_info['from_col'])}, 行 {img_info['from_row']}")
                                if img_info['to_col']:
                                    print(f"  结束位置: 列 {get_column_letter(img_info['to_col'])}, 行 {img_info['to_row']}")
                                    if img_info['from_col'] == img_info['to_col'] == column_num:
                                        print(f"  ✓✓ 图片完全在列 {column_letter} 内（行 {img_info['from_row']}-{img_info['to_row']}）")
                                    elif img_info['from_col'] <= column_num <= img_info['to_col']:
                                        print(f"  ⚠ 图片跨越多个列（包括列 {column_letter}）")
                                        print(f"    从列 {get_column_letter(img_info['from_col'])} 到列 {get_column_letter(img_info['to_col'])}")
                                    else:
                                        print(f"  ✗ 图片不在列 {column_letter} 范围内")
                                else:
                                    if img_info['from_col'] == column_num:
                                        print(f"  ✓✓ 图片在列 {column_letter} 内（行 {img_info['from_row']}）")
                                    else:
                                        print(f"  ✗ 图片不在列 {column_letter}（在列 {get_column_letter(img_info['from_col'])}）")
                        else:
                            print("  未能从drawing XML中提取图片位置信息")
                            
                except Exception as e:
                    print(f"\n检查关系时出错: {e}")
                    import traceback
                    traceback.print_exc()
        
        wb.close()
        return True
        
    except Exception as e:
        # 重新抛出异常，让调用者处理
        raise

if __name__ == "__main__":
    xls_file = "/Users/changun/Documents/新开模具清单.xls"
    xlsx_file = "/Users/changun/Documents/新开模具清单_temp.xlsx"
    
    # 首先尝试直接读取（可能是.xlsx格式但扩展名错误）
    print("尝试直接读取文件（可能是.xlsx格式）...\n")
    direct_read_success = False
    try:
        analyze_images_in_xlsx(xls_file, 'R')
        direct_read_success = True
        print("\n✓ 成功直接读取文件！")
    except Exception as e:
        error_msg = str(e)
        if "does not support the old .xls" in error_msg:
            print("✗ 确认这是旧的.xls格式文件\n")
            print("=" * 70)
            print("检测到.xls格式文件，尝试使用LibreOffice转换...")
            print("=" * 70)
            
            # 使用LibreOffice转换（保留图片）
            if convert_xls_to_xlsx_with_libreoffice(xls_file, xlsx_file):
                print("\n✓ 转换成功，正在分析图片...\n")
                analyze_images_in_xlsx(xlsx_file, 'R')
            else:
                print("\n自动转换失败，请手动转换文件：")
                print("1. 在Excel或Numbers中打开文件")
                print("2. 选择 '文件' -> '另存为' 或 '导出'")
                print("3. 将格式改为 '.xlsx' (Excel 2007+)")
                print("4. 保存后，将新文件路径提供给脚本")
        else:
            print(f"✗ 读取失败: {e}")
