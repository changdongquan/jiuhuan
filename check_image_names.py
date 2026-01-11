#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查Excel文件中图片的名称
"""
import zipfile
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter

def check_image_names(xlsx_path):
    """检查xlsx文件中图片的名称"""
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            # 查找drawing文件
            drawing_files = [f for f in zip_ref.namelist() if 'xl/drawings/drawing' in f and f.endswith('.xml')]
            if not drawing_files:
                print("未找到drawing文件")
                return
            
            drawing_path = drawing_files[0]
            
            # 读取drawing XML
            with zip_ref.open(drawing_path) as f:
                tree = ET.parse(f)
                root = tree.getroot()
            
            # 解析命名空间
            namespaces = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            
            image_count = 0
            named_count = 0
            
            # 检查所有锚点
            for anchor_type in ['twoCellAnchor', 'oneCellAnchor']:
                anchors = root.findall(f'.//xdr:{anchor_type}', namespaces)
                
                for idx, anchor in enumerate(anchors, 1):
                    image_count += 1
                    
                    # 获取位置信息
                    from_elem = anchor.find('.//xdr:from', namespaces)
                    col = None
                    row = None
                    
                    if from_elem is not None:
                        col_elem = from_elem.find('xdr:col', namespaces)
                        row_elem = from_elem.find('xdr:row', namespaces)
                        if col_elem is not None:
                            col = int(col_elem.text) + 1
                        if row_elem is not None:
                            row = int(row_elem.text) + 1
                    
                    # 查找图片名称
                    # 可能在以下位置：
                    # 1. cNvPr (非视觉绘图属性) - name属性
                    # 2. pic/cNvPr - name属性
                    # 3. spPr - 可能没有名称
                    
                    name = None
                    
                    # 检查pic元素中的cNvPr
                    pic = anchor.find('.//xdr:pic', namespaces)
                    if pic is not None:
                        cNvPr = pic.find('.//xdr:cNvPr', namespaces)
                        if cNvPr is not None:
                            name = cNvPr.get('name')
                    
                    # 如果没有找到，检查图形中的cNvPr
                    if name is None:
                        sp = anchor.find('.//xdr:sp', namespaces)
                        if sp is not None:
                            cNvPr = sp.find('.//xdr:cNvPr', namespaces)
                            if cNvPr is not None:
                                name = cNvPr.get('name')
                    
                    # 检查grpSp (组形状)
                    if name is None:
                        grpSp = anchor.find('.//xdr:grpSp', namespaces)
                        if grpSp is not None:
                            cNvPr = grpSp.find('.//xdr:cNvPr', namespaces)
                            if cNvPr is not None:
                                name = cNvPr.get('name')
                    
                    if name:
                        named_count += 1
                        if image_count <= 10 or name:  # 显示前10个和所有有名称的
                            col_letter = get_column_letter(col) if col else "?"
                            print(f"图片 {image_count}: 列 {col_letter}, 行 {row} → 名称: '{name}'")
                    else:
                        if image_count <= 10:  # 只显示前10个没有名称的
                            col_letter = get_column_letter(col) if col else "?"
                            print(f"图片 {image_count}: 列 {col_letter}, 行 {row} → 无名称")
            
            print(f"\n{'='*60}")
            print(f"统计信息：")
            print(f"  总图片数量：  {image_count} 个")
            print(f"  有名称的：    {named_count} 个 ({named_count/image_count*100:.1f}%)" if image_count > 0 else f"  有名称的：    {named_count} 个")
            print(f"  无名称的：    {image_count - named_count} 个 ({(image_count - named_count)/image_count*100:.1f}%)" if image_count > 0 else f"  无名称的：    {image_count - named_count} 个")
            
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    xlsx_file = "/Users/changun/Documents/新开模具清单.xlsx"
    print(f"检查文件: {xlsx_file}")
    print(f"{'='*60}\n")
    check_image_names(xlsx_file)

