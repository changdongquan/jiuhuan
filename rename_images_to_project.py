#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将图片名称改为所在行B列的项目编号
"""
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import shutil

def get_b_column_values(xlsx_path, sheet_name):
    """获取B列的所有值"""
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        
        b_column = {}
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)  # B列是第2列
            value = cell.value
            if value is not None and str(value).strip():
                b_column[row] = str(value).strip()
        
        wb.close()
        return b_column
        
    except Exception as e:
        print(f"读取B列错误: {e}")
        return {}

def rename_images_to_project_numbers(xlsx_path):
    """将图片名称改为所在行B列的项目编号"""
    try:
        # 获取B列值
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        sheet_name = wb.sheetnames[0]
        wb.close()
        
        b_column = get_b_column_values(xlsx_path, sheet_name)
        print(f"读取到 {len(b_column)} 行有项目编号")
        
        # 创建临时目录
        temp_dir = "/tmp/excel_rename_images"
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir)
        
        # 解压xlsx文件
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # 查找drawing文件
        drawing_files = []
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if 'xl/drawings/drawing' in os.path.join(root, file) and file.endswith('.xml'):
                    rel_path = os.path.relpath(os.path.join(root, file), temp_dir)
                    drawing_files.append(rel_path)
        
        if not drawing_files:
            print("未找到drawing文件")
            return False
        
        namespaces = {
            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        }
        
        renamed_count = 0
        skipped_count = 0
        
        for drawing_path in drawing_files:
            drawing_file = os.path.join(temp_dir, drawing_path)
            
            # 解析XML
            tree = ET.parse(drawing_file)
            root = tree.getroot()
            
            # 查找所有锚点
            for anchor_type in ['twoCellAnchor', 'oneCellAnchor']:
                anchors = root.findall(f'.//xdr:{anchor_type}', namespaces)
                
                for anchor in anchors:
                    # 获取图片位置
                    from_elem = anchor.find('.//xdr:from', namespaces)
                    if from_elem is None:
                        continue
                    
                    col_elem = from_elem.find('xdr:col', namespaces)
                    row_elem = from_elem.find('xdr:row', namespaces)
                    
                    if col_elem is None or row_elem is None:
                        continue
                    
                    row = int(row_elem.text) + 1  # Excel从0开始，转换为从1开始
                    
                    # 获取B列的项目编号
                    project_num = b_column.get(row, None)
                    
                    if not project_num:
                        skipped_count += 1
                        continue
                    
                    # 查找图片元素并更新名称
                    updated = False
                    
                    # 检查pic元素
                    pic = anchor.find('.//xdr:pic', namespaces)
                    if pic is not None:
                        cNvPr = pic.find('.//xdr:cNvPr', namespaces)
                        if cNvPr is not None:
                            old_name = cNvPr.get('name', '')
                            cNvPr.set('name', project_num)
                            renamed_count += 1
                            updated = True
                            if renamed_count <= 5:  # 显示前5个
                                print(f"行 {row}: '{old_name}' → '{project_num}'")
                    
                    # 检查sp元素（图形）
                    if not updated:
                        sp = anchor.find('.//xdr:sp', namespaces)
                        if sp is not None:
                            cNvPr = sp.find('.//xdr:cNvPr', namespaces)
                            if cNvPr is not None:
                                old_name = cNvPr.get('name', '')
                                cNvPr.set('name', project_num)
                                renamed_count += 1
                                updated = True
                                if renamed_count <= 5:
                                    print(f"行 {row}: '{old_name}' → '{project_num}'")
            
            # 保存修改后的XML
            tree.write(drawing_file, encoding='utf-8', xml_declaration=True)
            print(f"已更新drawing文件: {drawing_path}")
        
        if renamed_count == 0:
            print("没有找到需要重命名的图片")
            shutil.rmtree(temp_dir)
            return False
        
        # 创建备份
        backup_file = xlsx_path + '.bak'
        if os.path.exists(backup_file):
            os.remove(backup_file)
        shutil.copy2(xlsx_path, backup_file)
        print(f"\n已创建备份文件: {backup_file}")
        
        # 重新打包xlsx文件
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as new_zip:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    new_zip.write(file_path, arcname)
        
        # 清理临时目录
        shutil.rmtree(temp_dir)
        
        print(f"\n✓ 成功重命名 {renamed_count} 个图片")
        if skipped_count > 0:
            print(f"⚠ 跳过 {skipped_count} 个图片（所在行没有项目编号）")
        
        return True
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        return False

if __name__ == "__main__":
    xlsx_file = "/Users/changun/Documents/新开模具清单.xlsx"
    
    print("=" * 70)
    print("将图片名称改为所在行B列的项目编号")
    print("=" * 70)
    print()
    
    if rename_images_to_project_numbers(xlsx_file):
        print("\n✓ 重命名完成！")
    else:
        print("\n✗ 重命名失败")

