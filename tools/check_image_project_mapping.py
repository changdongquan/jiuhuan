#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查图片与B列项目编号的对应关系
"""
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys

def get_image_positions(xlsx_path):
    """获取所有图片的位置信息"""
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            drawing_files = [f for f in zip_ref.namelist() if 'xl/drawings/drawing' in f and f.endswith('.xml')]
            if not drawing_files:
                return []
            
            drawing_path = drawing_files[0]
            with zip_ref.open(drawing_path) as f:
                tree = ET.parse(f)
                root = tree.getroot()
            
            namespaces = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            
            images = []
            
            # 检查twoCellAnchor和oneCellAnchor
            for anchor_type in ['twoCellAnchor', 'oneCellAnchor']:
                for anchor in root.findall(f'.//xdr:{anchor_type}', namespaces):
                    from_elem = anchor.find('.//xdr:from', namespaces)
                    if from_elem is None:
                        continue
                    
                    col_elem = from_elem.find('xdr:col', namespaces)
                    row_elem = from_elem.find('xdr:row', namespaces)
                    
                    if col_elem is None or row_elem is None:
                        continue
                    
                    col = int(col_elem.text) + 1
                    row = int(row_elem.text) + 1
                    
                    # 获取图片名称
                    name = None
                    pic = anchor.find('.//xdr:pic', namespaces)
                    if pic is not None:
                        cNvPr = pic.find('.//xdr:cNvPr', namespaces)
                        if cNvPr is not None:
                            name = cNvPr.get('name')
                    
                    images.append({
                        'row': row,
                        'col': col,
                        'name': name
                    })
            
            return images
            
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_b_column_values(xlsx_path, sheet_name):
    """获取B列的所有值"""
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        
        b_column = {}
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)  # B列是第2列
            value = cell.value
            if value is not None and str(value).strip():
                b_column[row] = str(value).strip()
        
        wb.close()
        return b_column
        
    except Exception as e:
        print(f"读取B列错误: {e}")
        import traceback
        traceback.print_exc()
        return {}

def check_image_project_mapping(xlsx_path):
    """检查图片与B列项目编号的对应关系"""
    print(f"检查文件: {xlsx_path}")
    print("=" * 70)
    
    # 获取所有图片位置
    print("\n正在获取图片位置信息...")
    images = get_image_positions(xlsx_path)
    print(f"找到 {len(images)} 个图片")
    
    # 获取B列值
    print("\n正在读取B列（项目编号）...")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_name = wb.sheetnames[0]  # 假设第一个工作表
    wb.close()
    
    b_column = get_b_column_values(xlsx_path, sheet_name)
    print(f"B列有 {len(b_column)} 行有值")
    
    # 分析对应关系
    print("\n" + "=" * 70)
    print("图片与B列项目编号对应关系分析：")
    print("=" * 70)
    
    # 按行排序图片
    images.sort(key=lambda x: x['row'])
    
    matched = 0
    no_match = 0
    multiple_images = 0
    rows_with_images = {}
    
    # 统计每行的图片数量
    for img in images:
        row = img['row']
        if row not in rows_with_images:
            rows_with_images[row] = []
        rows_with_images[row].append(img)
    
    # 检查对应关系
    print("\n【对应关系详情】")
    print("-" * 70)
    
    for row in sorted(rows_with_images.keys()):
        imgs = rows_with_images[row]
        project_num = b_column.get(row, None)
        
        if len(imgs) > 1:
            multiple_images += 1
        
        if project_num:
            matched += len(imgs)
            print(f"行 {row:3d}: B列项目编号='{project_num}' → {len(imgs)}个图片")
            for idx, img in enumerate(imgs, 1):
                name = img['name'] or '无名称'
                col_letter = get_column_letter(img['col'])
                print(f"          [{idx}] 图片名称: '{name}' (列{col_letter})")
        else:
            no_match += len(imgs)
            print(f"行 {row:3d}: B列无项目编号 → {len(imgs)}个图片")
            for idx, img in enumerate(imgs, 1):
                name = img['name'] or '无名称'
                col_letter = get_column_letter(img['col'])
                print(f"          [{idx}] 图片名称: '{name}' (列{col_letter})")
    
    # 统计信息
    print("\n" + "=" * 70)
    print("【统计信息】")
    print("-" * 70)
    print(f"总图片数量：        {len(images)} 个")
    print(f"有对应项目编号：    {matched} 个 ({matched/len(images)*100:.1f}%)" if images else "有对应项目编号：    0 个")
    print(f"无对应项目编号：    {no_match} 个 ({no_match/len(images)*100:.1f}%)" if images else "无对应项目编号：    0 个")
    print(f"有多个图片的行：    {multiple_images} 行")
    print(f"唯一图片行数：      {len(rows_with_images)} 行")
    
    # 检查一对一关系
    print("\n" + "=" * 70)
    print("【一一对应检查】")
    print("-" * 70)
    
    one_to_one = 0
    one_to_many = 0
    many_to_one = 0
    
    for row in sorted(rows_with_images.keys()):
        imgs = rows_with_images[row]
        project_num = b_column.get(row, None)
        
        if project_num:
            if len(imgs) == 1:
                one_to_one += 1
            else:
                one_to_many += 1
    
    # 检查是否有多个行使用相同的项目编号（many-to-one）
    project_to_rows = {}
    for row in sorted(rows_with_images.keys()):
        project_num = b_column.get(row, None)
        if project_num:
            if project_num not in project_to_rows:
                project_to_rows[project_num] = []
            project_to_rows[project_num].append(row)
    
    for project_num, rows in project_to_rows.items():
        if len(rows) > 1:
            many_to_one += len(rows)
    
    print(f"一对一关系：        {one_to_one} 行（1个项目编号对应1个图片）")
    print(f"一对多关系：        {one_to_many} 行（1个项目编号对应多个图片）")
    print(f"多对一关系：        {many_to_one} 行（多个行使用相同项目编号）")
    
    # 结论
    print("\n" + "=" * 70)
    print("【结论】")
    print("-" * 70)
    
    if one_to_one == len(rows_with_images) and many_to_one == 0:
        print("✓ 完美的一一对应关系！")
        print("  • 每个项目编号对应一个图片")
        print("  • 每个图片都有对应的项目编号")
        print("  • 没有重复的项目编号")
    elif one_to_one > 0:
        print("⚠ 部分对应关系")
        print(f"  • {one_to_one} 行是一对一关系")
        if one_to_many > 0:
            print(f"  • {one_to_many} 行是一对多关系（1个项目编号对应多个图片）")
        if many_to_one > 0:
            print(f"  • {many_to_one} 行是多对一关系（多个行使用相同项目编号）")
    else:
        print("✗ 无法建立一一对应关系")
    
    return {
        'total_images': len(images),
        'matched': matched,
        'no_match': no_match,
        'one_to_one': one_to_one,
        'one_to_many': one_to_many,
        'many_to_one': many_to_one
    }

if __name__ == "__main__":
    xlsx_file = "/Users/changun/Documents/新开模具清单.xlsx"
    check_image_project_mapping(xlsx_file)

